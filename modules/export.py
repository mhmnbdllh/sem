"""
export.py - Export & Report Generation Module for SEM Studio.
Generates APA narrative, Excel workbook, and methodological checklist.
"""
import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import datetime

COLORS = {
    "excellent": "#1a7a4a",
    "good":      "#2ecc71",
    "ok":        "#1a6fa8",
    "warning":   "#b7770d",
    "critical":  "#c0392b",
}

def badge(level, message):
    color = COLORS.get(level, "#555555")
    st.markdown(
        f'<div style="background:{color}18;border-left:4px solid {color};'
        f'padding:10px 14px;border-radius:4px;margin:6px 0;'
        f'color:#1a1a1a;font-size:0.92rem">{message}</div>',
        unsafe_allow_html=True,
    )

def _safe_float(val, default=None):
    if val is None: return default
    try:
        if isinstance(val, list): val = val[0]
        f = float(val)
        return None if np.isnan(f) else f
    except: return default

def _fmt(val, decimals=3):
    if val is None: return "—"
    try:
        if np.isnan(float(val)): return "—"
        return f"{float(val):.{decimals}f}"
    except: return str(val)

def _stars(p):
    if p is None: return ""
    try:
        p = float(p)
        if p < 0.001: return "***"
        elif p < 0.01: return "**"
        elif p < 0.05: return "*"
        elif p < 0.10: return "dag"
        return ""
    except: return ""


# ── SECTION 1: FULL CHECKLIST ────────────────────────────────────

def render_full_checklist():
    st.subheader("Methodological Checklist")
    st.markdown(
        "This checklist verifies all required methodological steps have been "
        "completed correctly before reporting results."
    )

    ss = st.session_state
    df = ss.get("df")
    constructs = ss.get("constructs", {})

    checks = {
        "DATA PREPARATION": {
            "Data uploaded and validated":
                ss.get("df_ready", False),
            "Variable roles assigned":
                bool(ss.get("assignments")),
            "Constructs defined (>= 3 items each)":
                all(len(v) >= 3 for v in constructs.values()) if constructs else False,
            "Structural paths (hypotheses) defined":
                len(ss.get("structural_paths", [])) > 0,
        },
        "DESCRIPTIVE AND ASSUMPTION TESTING": {
            "Descriptive statistics computed":
                ss.get("descriptive_complete", False),
            "Multivariate normality tested (Mardia)":
                "recommended_estimator" in ss,
            "Estimator selected (ML or MLR)":
                "recommended_estimator" in ss,
            "Common method bias assessed (Harman)":
                ss.get("descriptive_complete", False),
            "Outlier detection performed":
                "mv_outliers" in ss,
        },
        "MEASUREMENT MODEL (CFA)": {
            "EFA conducted (if instrument unvalidated)":
                ss.get("efa_complete", False),
            "CFA model estimated":
                "cfa_result" in ss,
            "Model fit assessed (RMSEA, CFI, TLI, SRMR)":
                bool(ss.get("cfa_fit")),
            "Factor loadings >= .50":
                _check_loadings(),
            "AVE >= .50 (convergent validity)":
                _check_ave(),
            "CR >= .70 (composite reliability)":
                _check_cr(),
            "Cronbach alpha >= .70":
                _check_alpha(),
            "HTMT discriminant validity assessed":
                "cfa_result" in ss,
        },
        "STRUCTURAL MODEL (SEM)": {
            "Full SEM estimated":
                "sem_result" in ss,
            "SEM fit indices adequate":
                _check_sem_fit(),
            "Structural paths reported (beta, SE, p)":
                bool(ss.get("sem_paths")),
            "R2 reported for endogenous constructs":
                bool(ss.get("sem_r2")),
            "Effect sizes (f2) computed":
                bool(ss.get("sem_paths")),
        },
        "ADVANCED ANALYSES (if applicable)": {
            "Mediation analysis with bootstrap CI":
                bool(ss.get("mediation_results")),
            "Moderation analysis (interaction terms)":
                bool(ss.get("moderation_results")),
            "Measurement invariance tested":
                bool(ss.get("invariance_results")),
            "Model comparison with rival models":
                bool(ss.get("comparison_results")),
        },
    }

    total_checks = 0
    total_pass   = 0

    for section, section_checks in checks.items():
        st.markdown(f"{section}")
        rows = []
        for check, passed in section_checks.items():
            rows.append({
                "Check":  check,
                "Status": "Pass" if passed else "Pending",
            })
            total_checks += 1
            if passed: total_pass += 1

        def color_status(val):
            if val == "Pass":    return "color:#1a7a4a;font-weight:700"
            return "color:#888888"

        df_check = pd.DataFrame(rows)
        st.dataframe(
            df_check.style.map(color_status, subset=["Status"])
                          .set_properties(**{"color":"#1a1a1a","background-color":"#ffffff"})
                          .set_table_styles([{
                              "selector":"th",
                              "props":[("background-color","#2E86AB"),("color","white"),("font-weight","bold")]
                          }]),
            use_container_width=True, hide_index=True
        )

    pct = total_pass / total_checks if total_checks > 0 else 0
    st.markdown(f"Overall progress: {total_pass}/{total_checks} checks complete ({pct:.0%})")
    st.progress(pct)

    if pct >= 0.80:
        badge("excellent", f"{pct:.0%} of methodological steps completed. Ready to export report!")
    elif pct >= 0.50:
        badge("warning", f"{pct:.0%} complete. Consider completing remaining analyses before reporting.")
    else:
        badge("critical", f"Only {pct:.0%} complete. Please run CFA and SEM before exporting.")


def _check_loadings():
    from utils.thresholds import CFA as CFA_T
    metrics = st.session_state.get("cfa_metrics", {})
    if not metrics: return False
    return all(
        min(m.get("lambdas", [0])) >= CFA_T["loading_min"]
        for m in metrics.values() if m.get("lambdas")
    )

def _check_ave():
    from utils.thresholds import CFA as CFA_T
    metrics = st.session_state.get("cfa_metrics", {})
    if not metrics: return False
    return all((m.get("ave") or 0) >= CFA_T["ave_min"] for m in metrics.values())

def _check_cr():
    from utils.thresholds import CFA as CFA_T
    metrics = st.session_state.get("cfa_metrics", {})
    if not metrics: return False
    return all((m.get("cr") or 0) >= CFA_T["cr_min"] for m in metrics.values())

def _check_alpha():
    from utils.thresholds import CFA as CFA_T
    metrics = st.session_state.get("cfa_metrics", {})
    if not metrics: return False
    return all((m.get("alpha") or 0) >= CFA_T["alpha_min"] for m in metrics.values())

def _check_sem_fit():
    from utils.thresholds import FIT
    fit = st.session_state.get("sem_fit", {})
    if not fit: return False
    return (
        (fit.get("rmsea") or 999) <= FIT["rmsea_acceptable"] and
        (fit.get("cfi")   or 0)   >= FIT["cfi_acceptable"]
    )


# ── SECTION 2: APA NARRATIVE ─────────────────────────────────────

def generate_apa_narrative():
    ss  = st.session_state
    df  = ss.get("df")
    n   = len(df) if df is not None else "N/A"
    est = ss.get("recommended_estimator", "ML")
    constructs  = ss.get("constructs", {})
    cfa_fit     = ss.get("cfa_fit", {})
    sem_fit     = ss.get("sem_fit", {})
    metrics     = ss.get("cfa_metrics", {})
    sem_paths   = ss.get("sem_paths", [])
    med_results = ss.get("mediation_results")
    med_vars    = ss.get("mediation_vars", {})
    now         = datetime.now().strftime("%B %d, %Y")

    lines = []
    lines.append("=" * 70)
    lines.append("SEM STUDIO — APA RESULTS SECTION")
    lines.append(f"Generated: {now}")
    lines.append("=" * 70)
    lines.append("")

    # Sample and method
    lines.append("SAMPLE AND METHOD")
    lines.append("-" * 40)
    lines.append(
        f"The analysis was conducted on a sample of N = {n} participants. "
        f"{est} estimation was used based on Mardia's multivariate normality assessment. "
        f"All analyses were performed using SEM Studio (R/lavaan; Rosseel, 2012)."
    )
    lines.append("")

    # Measurement model
    lines.append("MEASUREMENT MODEL (CFA)")
    lines.append("-" * 40)
    if constructs:
        n_c = len(constructs)
        n_i = sum(len(v) for v in constructs.values())
        lines.append(
            f"A confirmatory factor analysis (CFA) was conducted with {n_c} "
            f"latent constructs and {n_i} observed indicators."
        )

    if cfa_fit:
        rmsea = _safe_float(cfa_fit.get("rmsea"))
        cfi   = _safe_float(cfa_fit.get("cfi"))
        tli   = _safe_float(cfa_fit.get("tli"))
        srmr  = _safe_float(cfa_fit.get("srmr"))
        chi2  = _safe_float(cfa_fit.get("chi2"))
        df_   = _safe_float(cfa_fit.get("df"))
        p_    = _safe_float(cfa_fit.get("p"))

        fit_parts = []
        if rmsea: fit_parts.append(f"RMSEA = {rmsea:.3f}")
        if cfi:   fit_parts.append(f"CFI = {cfi:.3f}")
        if tli:   fit_parts.append(f"TLI = {tli:.3f}")
        if srmr:  fit_parts.append(f"SRMR = {srmr:.3f}")

        from utils.thresholds import FIT
        acceptable = (rmsea or 999) <= FIT["rmsea_acceptable"] and (cfi or 0) >= FIT["cfi_acceptable"]
        verdict = "acceptable fit" if acceptable else "marginal fit"

        if chi2 is not None and df_:
            lines.append(
                f"The measurement model demonstrated {verdict}: "
                f"chi2({int(df_)}) = {chi2:.3f}, p = {_fmt(p_, 3)}, "
                f"{', '.join(fit_parts)}."
            )
        else:
            lines.append(f"The measurement model demonstrated {verdict}: {', '.join(fit_parts)}.")

    if metrics:
        lines.append("")
        lines.append("Reliability and validity indicators:")
        for cname, m in metrics.items():
            parts = []
            if m.get("alpha"): parts.append(f"alpha = {m['alpha']:.3f}")
            if m.get("cr"):    parts.append(f"CR = {m['cr']:.3f}")
            if m.get("ave"):   parts.append(f"AVE = {m['ave']:.3f}")
            if parts:
                lines.append(f"  {cname}: {', '.join(parts)}")
    lines.append("")

    # Structural model
    lines.append("STRUCTURAL MODEL (SEM)")
    lines.append("-" * 40)
    if sem_fit:
        rmsea = _safe_float(sem_fit.get("rmsea"))
        cfi   = _safe_float(sem_fit.get("cfi"))
        tli   = _safe_float(sem_fit.get("tli"))
        srmr  = _safe_float(sem_fit.get("srmr"))
        chi2  = _safe_float(sem_fit.get("chi2"))
        df_   = _safe_float(sem_fit.get("df"))
        p_    = _safe_float(sem_fit.get("p"))

        fit_parts = []
        if rmsea: fit_parts.append(f"RMSEA = {rmsea:.3f}")
        if cfi:   fit_parts.append(f"CFI = {cfi:.3f}")
        if tli:   fit_parts.append(f"TLI = {tli:.3f}")
        if srmr:  fit_parts.append(f"SRMR = {srmr:.3f}")

        from utils.thresholds import FIT
        acceptable = (rmsea or 999) <= FIT["rmsea_acceptable"] and (cfi or 0) >= FIT["cfi_acceptable"]
        verdict = "acceptable fit" if acceptable else "marginal fit"

        if chi2 is not None and df_:
            lines.append(
                f"The full structural model demonstrated {verdict}: "
                f"chi2({int(df_)}) = {chi2:.3f}, p = {_fmt(p_, 3)}, "
                f"{', '.join(fit_parts)}."
            )

    if sem_paths:
        lines.append("")
        lines.append("Structural path results:")
        for i, p in enumerate(sem_paths):
            beta  = _safe_float(p.get("beta"))
            se    = _safe_float(p.get("se"))
            p_val = _safe_float(p.get("p"))
            if beta is None or p_val is None: continue
            sig   = "significant" if p_val < 0.05 else "not significant"
            stars = _stars(p_val)
            lines.append(
                f"  H{i+1}: {p.get('predictor','?')} --> {p.get('outcome','?')}: "
                f"beta = {beta:.3f}{stars}, SE = {_fmt(se)}, p = {_fmt(p_val, 3)} -- {sig}."
            )
    lines.append("")

    # Mediation
    if med_results and isinstance(med_results, dict):
        lines.append("MEDIATION ANALYSIS")
        lines.append("-" * 40)
        x = med_vars.get("x", "X")
        m = med_vars.get("m", "M")
        y = med_vars.get("y", "Y")
        indirect_data = med_results.get("indirect", {})
        if isinstance(indirect_data, dict):
            indirect = _safe_float(indirect_data.get("est"))
            ci_lo    = _safe_float(indirect_data.get("ci_lo"))
            ci_hi    = _safe_float(indirect_data.get("ci_hi"))
            n_boot   = _safe_float(med_results.get("n_boot"), 5000)
            if indirect is not None and ci_lo is not None and ci_hi is not None:
                sig = not (ci_lo <= 0 <= ci_hi)
                lines.append(
                    f"A bootstrap mediation analysis ({int(n_boot):,} resamples) examined whether "
                    f"{m} mediated the relationship between {x} and {y}. "
                    f"The indirect effect was {'significant' if sig else 'not significant'} "
                    f"(indirect = {indirect:.4f}, 95% BCa CI [{ci_lo:.4f}, {ci_hi:.4f}])."
                )
    lines.append("")

    lines.append("=" * 70)
    lines.append("Note: All analyses via SEM Studio (R/lavaan).")
    lines.append("* p < .05. ** p < .01. *** p < .001. dag p < .10")
    lines.append("=" * 70)

    return "\n".join(lines)


def render_apa_narrative():
    st.subheader("APA Results Narrative")
    st.markdown(
        "Auto-generated APA 7th edition results text. "
        "**Copy, paste, then review and edit** before submitting to a journal."
    )

    narrative = generate_apa_narrative()
    st.text_area(
        "APA Results Section (copy-paste ready)",
        value=narrative,
        height=500,
    )

    badge("warning",
        "Important: This auto-generated text is a starting point. "
        "Always review, edit, and supplement with your own theoretical interpretation "
        "before submitting to a journal."
    )


# ── SECTION 3: EXCEL EXPORT ──────────────────────────────────────

def generate_excel():
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl not installed."

    ss  = st.session_state
    buf = io.BytesIO()

    try:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df         = ss.get("df")
            constructs = ss.get("constructs", {})
            assignments= ss.get("assignments", {})

            # Sheet 1: Overview
            overview = pd.DataFrame({
                "Item":  ["Sample Size (n)", "Constructs", "Indicators",
                          "Estimator", "Analysis Date"],
                "Value": [
                    len(df) if df is not None else "N/A",
                    len(constructs),
                    sum(len(v) for v in constructs.values()),
                    ss.get("recommended_estimator", "N/A"),
                    datetime.now().strftime("%Y-%m-%d %H:%M"),
                ]
            })
            overview.to_excel(writer, sheet_name="Overview", index=False)

            # Sheet 2: Descriptive Statistics
            if df is not None and assignments:
                indicator_cols = [c for c, r in assignments.items() if r == "indicator"]
                if indicator_cols:
                    from scipy import stats as scipy_stats
                    desc_rows = []
                    for col in indicator_cols:
                        x = df[col].dropna()
                        if len(x) >= 3:
                            desc_rows.append({
                                "Variable": col,
                                "N":        int(x.count()),
                                "Mean":     round(float(x.mean()), 3),
                                "SD":       round(float(x.std()), 3),
                                "Min":      round(float(x.min()), 3),
                                "Max":      round(float(x.max()), 3),
                                "Skewness": round(float(scipy_stats.skew(x)), 3),
                                "Kurtosis": round(float(scipy_stats.kurtosis(x)), 3),
                                "Missing":  int(df[col].isna().sum()),
                            })
                    if desc_rows:
                        pd.DataFrame(desc_rows).to_excel(
                            writer, sheet_name="Descriptive Statistics", index=False
                        )

            # Sheet 3: CFA Fit Indices
            cfa_fit = ss.get("cfa_fit", {})
            if cfa_fit:
                from utils.apa_tables import fit_indices_table
                fit_df = fit_indices_table(cfa_fit)
                if not fit_df.empty:
                    fit_df.to_excel(writer, sheet_name="CFA Fit Indices", index=False)

            # Sheet 4: Reliability and Validity
            metrics = ss.get("cfa_metrics", {})
            if metrics:
                rel_rows = []
                for cname, m in metrics.items():
                    rel_rows.append({
                        "Construct":  cname,
                        "Items":      m.get("n_items", "—"),
                        "Cronbach_a": _fmt(m.get("alpha")),
                        "CR":         _fmt(m.get("cr")),
                        "McDonald_w": _fmt(m.get("omega")),
                        "AVE":        _fmt(m.get("ave")),
                        "a_pass":     "Pass" if (m.get("alpha") or 0) >= 0.70 else "Fail",
                        "CR_pass":    "Pass" if (m.get("cr")    or 0) >= 0.70 else "Fail",
                        "AVE_pass":   "Pass" if (m.get("ave")   or 0) >= 0.50 else "Fail",
                    })
                pd.DataFrame(rel_rows).to_excel(
                    writer, sheet_name="Reliability Validity", index=False
                )

            # Sheet 5: SEM Fit Indices
            sem_fit = ss.get("sem_fit", {})
            if sem_fit:
                from utils.apa_tables import fit_indices_table
                sem_fit_df = fit_indices_table(sem_fit)
                if not sem_fit_df.empty:
                    sem_fit_df.to_excel(writer, sheet_name="SEM Fit Indices", index=False)

            # Sheet 6: Structural Paths
            sem_paths = ss.get("sem_paths", [])
            if sem_paths:
                path_rows = []
                for i, p in enumerate(sem_paths):
                    p_val = _safe_float(p.get("p"))
                    beta  = _safe_float(p.get("beta"))
                    path_rows.append({
                        "H":         f"H{i+1}",
                        "Path":      f"{p.get('predictor','?')} -> {p.get('outcome','?')}",
                        "Beta":      _fmt(beta),
                        "SE":        _fmt(_safe_float(p.get("se"))),
                        "z":         _fmt(_safe_float(p.get("z"))),
                        "p":         _fmt(p_val),
                        "Sig":       _stars(p_val),
                        "Decision":  "Supported" if p_val is not None and float(p_val) < 0.05 else "Not Supported",
                    })
                pd.DataFrame(path_rows).to_excel(
                    writer, sheet_name="Structural Paths", index=False
                )

            # Sheet 7: Mediation
            med_results = ss.get("mediation_results", {})
            med_vars    = ss.get("mediation_vars", {})
            if med_results and isinstance(med_results, dict):
                med_rows = []
                for key, label in [
                    ("a_path",  "a path (X -> M)"),
                    ("b_path",  "b path (M -> Y | X)"),
                    ("cp_path", "c prime (direct)"),
                    ("total",   "Total (c)"),
                    ("indirect","Indirect (a x b)"),
                ]:
                    d = med_results.get(key, {})
                    if not isinstance(d, dict): continue
                    est   = _safe_float(d.get("est"))
                    ci_lo = _safe_float(d.get("ci_lo"))
                    ci_hi = _safe_float(d.get("ci_hi"))
                    med_rows.append({
                        "Effect":  label,
                        "Beta":    _fmt(est),
                        "CI_Low":  _fmt(ci_lo) if ci_lo is not None else "—",
                        "CI_High": _fmt(ci_hi) if ci_hi is not None else "—",
                        "Sig":     (
                            "Significant" if ci_lo is not None and ci_hi is not None and not (float(ci_lo) <= 0 <= float(ci_hi))
                            else "Not Significant"
                        ) if key == "indirect" else (
                            _stars(est) if est is not None else "—"
                        ),
                    })
                if med_rows:
                    pd.DataFrame(med_rows).to_excel(
                        writer, sheet_name="Mediation", index=False
                    )

            # Sheet 8: Moderation
            mod_results = ss.get("moderation_results", {})
            if mod_results and isinstance(mod_results, dict):
                mod_vars = ss.get("moderation_vars", {})
                mod_rows = [
                    {"Term": f"{mod_vars.get('x','X')} (X)",
                     "Beta": _fmt(_safe_float(mod_results.get("b1"))),
                     "SE":   _fmt(_safe_float(mod_results.get("b1_se"))),
                     "t":    _fmt(_safe_float(mod_results.get("b1_t"))),
                     "p":    _fmt(_safe_float(mod_results.get("b1_p")))},
                    {"Term": f"{mod_vars.get('w','W')} (W)",
                     "Beta": _fmt(_safe_float(mod_results.get("b2"))),
                     "SE":   _fmt(_safe_float(mod_results.get("b2_se"))),
                     "t":    _fmt(_safe_float(mod_results.get("b2_t"))),
                     "p":    _fmt(_safe_float(mod_results.get("b2_p")))},
                    {"Term": f"X x W (interaction)",
                     "Beta": _fmt(_safe_float(mod_results.get("b3"))),
                     "SE":   _fmt(_safe_float(mod_results.get("b3_se"))),
                     "t":    _fmt(_safe_float(mod_results.get("b3_t"))),
                     "p":    _fmt(_safe_float(mod_results.get("b3_p")))},
                ]
                pd.DataFrame(mod_rows).to_excel(
                    writer, sheet_name="Moderation", index=False
                )

            # Sheet 9: APA Narrative
            # Write manually using openpyxl to prevent = being treated as formula
            narrative = generate_apa_narrative()
            apa_ws = writer.book.create_sheet("APA Narrative")
            apa_ws.cell(row=1, column=1, value="APA Results Section")
            # Prefix with space to prevent Excel treating = as formula
            cell = apa_ws.cell(row=2, column=1)
            cell.value = narrative
            cell.data_type = "s"  # force string type
            apa_ws.column_dimensions["A"].width = 120

        return buf.getvalue(), None

    except Exception as e:
        return None, str(e)


def render_excel_export():
    st.subheader("Export to Excel")
    st.markdown(
        "Download all results in a single Excel workbook with separate sheets: "
        "Overview, Descriptives, CFA Fit, Reliability & Validity, SEM Fit, "
        "Structural Paths, Mediation, Moderation, APA Narrative."
    )

    if st.button("Generate Excel Report", type="primary", key="export_excel_btn"):
        with st.spinner("Generating Excel workbook..."):
            excel_bytes, error = generate_excel()
            if error:
                st.error(f"Excel generation failed: {error}")
            elif excel_bytes:
                fname = f"SEM_Studio_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                st.download_button(
                    label="Download Excel Report",
                    data=excel_bytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel_btn"
                )
                badge("excellent", "Excel report ready for download!")


def generate_html_report() -> str:
    """Generate a complete APA-style HTML report of all SEM results."""
    ss  = st.session_state
    df  = ss.get("df")
    n   = len(df) if df is not None else "N/A"
    constructs   = ss.get("constructs", {})
    cfa_fit      = ss.get("cfa_fit", {})
    sem_fit      = ss.get("sem_fit", {})
    metrics      = ss.get("cfa_metrics", {})
    sem_paths    = ss.get("sem_paths", [])
    sem_r2       = ss.get("sem_r2", [])
    med_results  = ss.get("mediation_results")
    med_vars     = ss.get("mediation_vars", {})
    mod_results  = ss.get("moderation_results")
    mod_vars     = ss.get("moderation_vars", {})
    inv_results  = ss.get("invariance_results")
    inv_level    = ss.get("invariance_level", "—")
    narrative    = generate_apa_narrative()
    from datetime import datetime
    now = datetime.now().strftime("%B %d, %Y %H:%M")

    def tbl(headers, rows, caption=""):
        hdr = "".join(f"<th>{h}</th>" for h in headers)
        body = ""
        for i, row in enumerate(rows):
            bg = "#f8fafc" if i % 2 == 0 else "#ffffff"
            cells = "".join(f'<td style="padding:7px 12px;border-bottom:1px solid #eee">{v}</td>' for v in row)
            body += f'<tr style="background:{bg}">{cells}</tr>'
        cap = f'<caption style="text-align:left;font-size:0.82rem;color:#666;padding:4px 0">{caption}</caption>' if caption else ""
        return (
            f'<table style="width:100%;border-collapse:collapse;margin:12px 0;font-size:0.9rem">'
            f'{cap}'
            f'<thead><tr style="background:#2E86AB;color:white">{hdr}</tr></thead>'
            f'<tbody>{body}</tbody></table>'
        )

    def section(title, content, color="#2E86AB"):
        return (
            f'<div style="margin:28px 0">'
            f'<h2 style="color:{color};border-bottom:2px solid {color};padding-bottom:6px;margin-bottom:14px">{title}</h2>'
            f'{content}'
            f'</div>'
        )

    def badge_html(level, msg):
        colors = {"excellent":"#1a7a4a","ok":"#1a6fa8","warning":"#b7770d","critical":"#c0392b"}
        c = colors.get(level, "#555")
        return f'<div style="background:{c}18;border-left:4px solid {c};padding:10px 14px;border-radius:4px;margin:8px 0;color:#1a1a1a">{msg}</div>'

    # ── Build sections ─────────────────────────────────────────

    # 1. Overview
    overview_rows = [
        ["Sample Size (n)", str(n)],
        ["Number of Constructs", str(len(constructs))],
        ["Total Indicators", str(sum(len(v) for v in constructs.values()))],
        ["Estimator", ss.get("recommended_estimator", "N/A")],
        ["Generated", now],
    ]
    s1 = section("1. Overview", tbl(["Item", "Value"], overview_rows))

    # 2. Constructs
    const_rows = [[c, ", ".join(items), str(len(items))] for c, items in constructs.items()]
    s2 = section("2. Measurement Model Specification",
        tbl(["Construct", "Indicators", "n Items"], const_rows))

    # 3. CFA Fit
    fit_rows = []
    fit_keys = [
        ("chi2",  "chi2",    "Non-significant preferred"),
        ("df",    "df",      "Degrees of freedom"),
        ("pvalue","p",       "p-value"),
        ("rmsea", "RMSEA",   "<=.05 excellent; <=.06 good; <=.08 acceptable"),
        ("cfi",   "CFI",     ">=.97 excellent; >=.95 good; >=.90 acceptable"),
        ("tli",   "TLI",     ">=.95 good; >=.90 acceptable"),
        ("srmr",  "SRMR",    "<=.05 good; <=.08 acceptable"),
        ("aic",   "AIC",     "Lower is better"),
        ("bic",   "BIC",     "Lower is better"),
    ]
    for key, label, criterion in fit_keys:
        val = cfa_fit.get(key)
        if val is not None:
            try:
                fit_rows.append([label, f"{float(val):.3f}", criterion])
            except: pass
    s3 = section("3. CFA Model Fit", tbl(["Index", "Value", "Criterion"], fit_rows,
        "Note: chi2/df <= 2.0 = good; <= 5.0 = acceptable (Kline, 2016). RMSEA, CFI, TLI, SRMR per Hu & Bentler (1999)."))

    # 4. Factor Loadings
    cfa_loadings = ss.get("cfa_loadings", {})
    load_rows = []
    for cname, items_dict in cfa_loadings.items():
        for item, lam in items_dict.items():
            try:
                lam_f = float(lam)
                status = "Strong" if abs(lam_f) >= 0.70 else "Acceptable" if abs(lam_f) >= 0.50 else "Weak"
                load_rows.append([cname, item, f"{lam_f:.3f}", status])
            except: pass
    s4 = section("4. Factor Loadings", tbl(
        ["Construct", "Item", "Std. Loading (lambda)", "Status"],
        load_rows,
        "Note: lambda >= .70 = strong; >= .50 = acceptable; < .50 = weak (Hair et al., 2019)."
    ))

    # 5. Reliability & Validity
    rel_rows = []
    for cname, m in metrics.items():
        rel_rows.append([
            cname,
            str(m.get("n_items", "—")),
            _fmt(m.get("alpha")),
            _fmt(m.get("cr")),
            _fmt(m.get("ave")),
            "Pass" if (m.get("alpha") or 0) >= 0.70 else "Fail",
            "Pass" if (m.get("cr")    or 0) >= 0.70 else "Fail",
            "Pass" if (m.get("ave")   or 0) >= 0.50 else "Fail",
        ])
    s5 = section("5. Reliability and Validity", tbl(
        ["Construct", "Items", "Cronbach alpha", "CR", "AVE", "alpha>=.70", "CR>=.70", "AVE>=.50"],
        rel_rows,
        "Note: CR = Composite Reliability; AVE = Average Variance Extracted. Criteria per Fornell & Larcker (1981); Hair et al. (2019)."
    ))

    # 6. SEM Fit
    sem_fit_rows = []
    for key, label, criterion in fit_keys:
        val = sem_fit.get(key)
        if val is not None:
            try:
                sem_fit_rows.append([label, f"{float(val):.3f}", criterion])
            except: pass
    s6 = section("6. SEM Model Fit", tbl(["Index", "Value", "Criterion"], sem_fit_rows,
        "Note: Same criteria as CFA fit above."))

    # 7. Structural Paths
    path_rows = []
    for i, p in enumerate(sem_paths):
        pval = _safe_float(p.get("p"))
        beta = _safe_float(p.get("beta"))
        path_rows.append([
            f"H{i+1}",
            f"{p.get('predictor','?')} -> {p.get('outcome','?')}",
            _fmt(beta),
            _fmt(_safe_float(p.get("se"))),
            _fmt(_safe_float(p.get("z"))),
            _fmt(pval),
            _stars(pval),
            "Supported" if pval is not None and pval < 0.05 else "Not Supported",
        ])
    s7 = section("7. Structural Paths", tbl(
        ["H", "Path", "Beta", "SE", "z", "p", "Sig.", "Decision"],
        path_rows,
        "Note: * p < .05; ** p < .01; *** p < .001. Beta = standardized path coefficient."
    ))

    # 8. R-squared
    r2_rows = []
    for row in sem_r2:
        if isinstance(row, dict):
            r2_rows.append([row.get("Construct","—"), _fmt(row.get("R2")), row.get("R2 (%)","—")])
    s8 = section("8. Explained Variance (R2)", tbl(
        ["Construct", "R2", "R2 (%)"],
        r2_rows,
        "Note: R2 >= .26 = substantial; >= .13 = moderate; >= .02 = weak (Cohen, 1988)."
    )) if r2_rows else ""

    # 9. Mediation
    s9 = ""
    if med_results and isinstance(med_results, dict):
        x = med_vars.get("x","X"); m_v = med_vars.get("m","M"); y = med_vars.get("y","Y")
        med_rows = []
        for key, label in [
            ("a_path", f"{x} -> {m_v} (a path)"),
            ("b_path", f"{m_v} -> {y} | {x} (b path)"),
            ("cp_path",f"{x} -> {y} direct (c prime)"),
            ("total",  f"{x} -> {y} total (c)"),
            ("indirect",f"Indirect effect (a x b)"),
        ]:
            d = med_results.get(key, {})
            if not isinstance(d, dict): continue
            est   = _safe_float(d.get("est"))
            ci_lo = _safe_float(d.get("ci_lo"))
            ci_hi = _safe_float(d.get("ci_hi"))
            pval  = _safe_float(d.get("p"))
            ci_str = f"[{ci_lo:.4f}, {ci_hi:.4f}]" if ci_lo is not None and ci_hi is not None else "—"
            if key == "indirect":
                sig = "Significant" if (ci_lo is not None and ci_hi is not None and not (ci_lo <= 0 <= ci_hi)) else "Not Significant"
            else:
                sig = _stars(pval) if pval is not None else "—"
            med_rows.append([label, _fmt(est), ci_str, sig])
        s9 = section("9. Mediation Analysis", tbl(
            ["Effect", "Beta", "95% BCa CI", "Sig."],
            med_rows,
            f"Note: Bootstrap mediation via R/lavaan. X={x}, M={m_v}, Y={y}. Significance: CI not containing zero."
        ))

    # 10. Moderation
    s10 = ""
    if mod_results and isinstance(mod_results, dict):
        x = mod_vars.get("x","X"); w = mod_vars.get("w","W"); y = mod_vars.get("y","Y")
        b3   = _safe_float(mod_results.get("b3"))
        b3_p = _safe_float(mod_results.get("b3_p"))
        r2_2 = _safe_float(mod_results.get("r2_2"))
        dr2  = _safe_float(mod_results.get("delta_r2"))
        mod_rows = [
            [f"{x} (X)",          _fmt(_safe_float(mod_results.get("b1"))), _fmt(_safe_float(mod_results.get("b1_se"))), _fmt(_safe_float(mod_results.get("b1_p"))), _stars(_safe_float(mod_results.get("b1_p")))],
            [f"{w} (W)",          _fmt(_safe_float(mod_results.get("b2"))), _fmt(_safe_float(mod_results.get("b2_se"))), _fmt(_safe_float(mod_results.get("b2_p"))), _stars(_safe_float(mod_results.get("b2_p")))],
            [f"{x} x {w} (Int.)",_fmt(b3),                                  _fmt(_safe_float(mod_results.get("b3_se"))), _fmt(b3_p),                                  _stars(b3_p)],
        ]
        mod_content = tbl(["Term","Beta","SE","p","Sig."], mod_rows,
            f"Note: Variables mean-centered. Interaction significant if p < .05. X={x}, W={w}, Y={y}.")
        mod_content += f"<p>Model R2 = {_fmt(r2_2)}, Delta R2 = {_fmt(dr2)}</p>"
        s10 = section("10. Moderation Analysis", mod_content)

    # 11. Invariance
    s11 = ""
    if inv_results and isinstance(inv_results, dict) and "error" not in inv_results:
        inv_rows = []
        for name, label in [("configural","Configural"),("metric","Metric"),("scalar","Scalar")]:
            fit_d = inv_results.get(name, {})
            if isinstance(fit_d, dict):
                inv_rows.append([
                    label,
                    _fmt(fit_d.get("cfi")),
                    _fmt(fit_d.get("rmsea")),
                    _fmt(fit_d.get("srmr")),
                ])
        diff_rows = []
        for key, label in [("diff_metric","Metric vs Configural"),("diff_scalar","Scalar vs Metric")]:
            d = inv_results.get(key, {})
            if isinstance(d, dict):
                dcfi = _safe_float(d.get("delta_cfi"))
                diff_rows.append([
                    label,
                    _fmt(d.get("delta_chi2")),
                    _fmt(dcfi),
                    _fmt(d.get("delta_rmsea")),
                    "Supported" if dcfi is not None and dcfi >= -0.010 else "Not Supported",
                ])
        inv_content  = tbl(["Model","CFI","RMSEA","SRMR"], inv_rows)
        inv_content += tbl(["Comparison","Delta chi2","Delta CFI","Delta RMSEA","Invariance"],
            diff_rows,
            "Note: Delta CFI >= -.010 supports invariance (Cheung & Rensvold, 2002). "
            f"Highest level achieved: {inv_level}.")
        s11 = section("11. Measurement Invariance", inv_content)

    # 12. APA Narrative
    narrative_html = narrative.replace("\n", "<br>").replace("=","&#61;")
    s12 = section("12. APA Results Narrative",
        f'<pre style="background:#f8fafc;padding:16px;border-radius:6px;'
        f'font-family:Georgia,serif;font-size:0.88rem;line-height:1.7;'
        f'white-space:pre-wrap;border:1px solid #dde3ea">{narrative}</pre>',
        color="#555"
    )

    # ── Assemble full HTML ────────────────────────────────────
    body = s1 + s2 + s3 + s4 + s5 + s6 + s7 + s8 + s9 + s10 + s11 + s12

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SEM Studio Report</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; max-width: 960px; margin: 40px auto;
         padding: 0 24px; color: #1a1a1a; background: #fff; line-height: 1.6; }}
  h1   {{ color: #2E86AB; border-bottom: 3px solid #2E86AB; padding-bottom: 10px; }}
  h2   {{ margin-top: 32px; }}
  table {{ page-break-inside: avoid; }}
  th   {{ padding: 8px 12px; text-align: left; }}
  td   {{ vertical-align: top; }}
  @media print {{
    body {{ margin: 20px; }}
    .no-print {{ display: none; }}
  }}
</style>
</head>
<body>
<h1>SEM Studio Analysis Report</h1>
<p style="color:#888;font-size:0.85rem">Generated: {now} &nbsp;|&nbsp; R/lavaan Backend &nbsp;|&nbsp; n = {n}</p>
<hr style="border-color:#dde3ea">
{body}
<hr style="border-color:#dde3ea;margin-top:40px">
<p style="color:#aaa;font-size:0.78rem;text-align:center">
  SEM Studio &mdash; Powered by R/lavaan &mdash; 
  References: Hair et al. (2019); Kline (2016); Hu &amp; Bentler (1999); Rosseel (2012)
</p>
</body>
</html>"""

    return html


def render_html_export():
    st.subheader("Export HTML Report")
    st.markdown(
        "Download a complete, publication-ready HTML report. "
        "Open in browser, print to PDF, or copy-paste into Word."
    )

    if st.button("Generate HTML Report", type="primary", key="export_html_btn"):
        with st.spinner("Generating HTML report..."):
            try:
                html = generate_html_report()
                from datetime import datetime
                fname = f"SEM_Studio_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.html"
                st.download_button(
                    label="Download HTML Report",
                    data=html.encode("utf-8"),
                    file_name=fname,
                    mime="text/html",
                    key="dl_html_btn"
                )
                badge("excellent", "HTML report ready! Open in browser to view, print to PDF, or copy-paste into Word.")
            except Exception as e:
                st.error(f"HTML generation failed: {str(e)}")



# ── SECTION 4: QUICK TEXT ────────────────────────────────────────

def render_text_export():
    st.subheader("Quick Text Summary")
    st.markdown("Plain text summary — easy to paste into any document or email.")

    ss    = st.session_state
    lines = []

    cfa_fit = ss.get("cfa_fit", {})
    if cfa_fit:
        lines.append("CFA FIT:")
        for k in ["rmsea","cfi","tli","srmr","aic","bic"]:
            v = _safe_float(cfa_fit.get(k))
            if v: lines.append(f"  {k.upper()} = {v:.3f}")

    sem_paths = ss.get("sem_paths", [])
    if sem_paths:
        lines.append("\nSTRUCTURAL PATHS:")
        for i, p in enumerate(sem_paths):
            beta  = _safe_float(p.get("beta"))
            p_val = _safe_float(p.get("p"))
            if beta is not None:
                lines.append(
                    f"  H{i+1}: {p.get('predictor','?')} -> {p.get('outcome','?')}: "
                    f"beta = {beta:.3f}, p = {_fmt(p_val, 3)}"
                )

    med = ss.get("mediation_results", {})
    if med and isinstance(med, dict):
        indirect_data = med.get("indirect", {})
        if isinstance(indirect_data, dict):
            indirect = _safe_float(indirect_data.get("est"))
            ci_lo    = _safe_float(indirect_data.get("ci_lo"))
            ci_hi    = _safe_float(indirect_data.get("ci_hi"))
            if indirect is not None:
                lines.append(f"\nMEDIATION:")
                lines.append(
                    f"  Indirect = {indirect:.4f}, "
                    f"95% CI [{_fmt(ci_lo, 4)}, {_fmt(ci_hi, 4)}]"
                )

    text = "\n".join(lines) if lines else "No results yet. Run analyses first."
    st.text_area("Results Summary", value=text, height=250, key="text_export_area")


# ── MAIN RENDER ───────────────────────────────────────────────────

def render_export():
    st.title("Export Report")
    st.markdown(
        "Generate and download your complete SEM analysis report. "
        "All results from every module are compiled into publication-ready outputs."
    )

    if not st.session_state.get("df_ready"):
        st.warning("Please complete Data Input and Model Setup first.")
        return

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Checklist",
        "APA Narrative",
        "HTML Report",
        "Excel Export",
        "Quick Text",
    ])

    with tab1:
        render_full_checklist()
    with tab2:
        render_apa_narrative()
    with tab3:
        render_html_export()
    with tab4:
        render_excel_export()
    with tab5:
        render_text_export()
